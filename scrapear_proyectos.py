import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# Función para obtener el contacto del arquitecto desde su página
def get_architect_contact(architect_url):
    try:
        print(f"Buscando contacto del arquitecto en: {architect_url}")
        response = requests.get(architect_url, headers={"User-Agent": "Mozilla/5.0"})
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        # Buscar el enlace de contacto dentro de la clase específica
        contact = soup.find('a', class_='js-office-website')
        if contact:
            return contact['href'].strip()
        return "No disponible"
    except Exception as e:
        print(f"Error al buscar contacto en {architect_url}: {e}")
        return "No disponible"

# Función para extraer información de un proyecto
def scrape_project(url):
    print(f"Extrayendo datos de: {url}")
    try:
        response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        project_data = {
            'Título': None,
            'Arquitecto': None,
            'Link Arquitecto': None,
            'Contacto Arquitecto': None,
            'Ubicación': None,
            'Año': None,
            'Fotógrafo': None,
            'Descripción': None,
            'Link Proyecto': url
        }

        # Título
        title = soup.find('h1', class_='afd-title-big')
        project_data['Título'] = title.text.strip() if title else "No disponible"
        print(f"Título: {project_data['Título']}")

        # Arquitectos
        specs_list = soup.find_all('li', class_='afd-specs__item')
        for spec in specs_list:
            key = spec.find('span', class_='afd-specs__key')
            value = spec.find('span', class_='afd-specs__value')
            if key and value:
                key_text = key.text.strip()
                if "Architects:" in key_text:
                    project_data['Arquitecto'] = value.text.strip()
                    link = value.find('a')
                    project_data['Link Arquitecto'] = link['href'] if link else "No disponible"
        print(f"Arquitecto: {project_data['Arquitecto']}")
        print(f"Link Arquitecto: {project_data['Link Arquitecto']}")

        # Obtener el contacto del arquitecto
        if project_data['Link Arquitecto'] and project_data['Link Arquitecto'] != "No disponible":
            project_data['Contacto Arquitecto'] = get_architect_contact(project_data['Link Arquitecto'])
        print(f"Contacto Arquitecto: {project_data['Contacto Arquitecto']}")

        # Ubicación
        location = soup.find('div', class_='afd-specs__header-location')
        project_data['Ubicación'] = location.text.strip() if location else "No disponible"
        print(f"Ubicación: {project_data['Ubicación']}")

        # Año
        year_item = next((item for item in specs_list if "Year" in (item.text or "")), None)
        if year_item:
            year_value = year_item.find('span', class_='afd-specs__value')
            project_data['Año'] = year_value.text.strip() if year_value else "No disponible"
        else:
            project_data['Año'] = "No disponible"
        print(f"Año: {project_data['Año']}")

        # Fotógrafos
        photographers = soup.find('div', class_='afd-specs__photographers')
        if photographers:
            project_data['Fotógrafo'] = photographers.find('span', class_='afd-specs__value').text.strip()
        else:
            project_data['Fotógrafo'] = "No disponible"
        print(f"Fotógrafo: {project_data['Fotógrafo']}")

        # Descripción
        description = soup.find('div', class_='afd-article-content')
        project_data['Descripción'] = description.text.strip() if description else "No disponible"
        print(f"Descripción: {project_data['Descripción'][:100]}...")  # Mostrar primeros 100 caracteres

        return project_data

    except Exception as e:
        print(f"Error al procesar {url}: {e}")
        return None

# Función para personalizar el Excel
def customize_excel(file_name):
    wb = load_workbook(file_name)
    ws = wb.active

    # Agregar columnas adicionales
    ws['I1'] = 'Contactado'
    ws['J1'] = 'Fecha del Contacto'
    ws['K1'] = 'Estado'

    # Crear listas desplegables
    contactado_validation = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
    estado_validation = DataValidation(type="list", formula1='"Contactado,Esperando,Respondido,Denegado"', allow_blank=True)

    ws.add_data_validation(contactado_validation)
    ws.add_data_validation(estado_validation)

    contactado_validation.add(f"I2:I{ws.max_row}")
    estado_validation.add(f"K2:K{ws.max_row}")

    # Guardar el archivo
    wb.save(file_name)
    print(f"Archivo Excel personalizado: {file_name}")

# Función principal para procesar los enlaces desde múltiples archivos
def main():
    try:
        import glob
        link_files = glob.glob("*_links.txt")

        for link_file in link_files:
            categoria = link_file.replace("_links.txt", "")
            print(f"\n📂 Procesando categoría: {categoria}")

            with open(link_file, "r") as f:
                project_links = f.read().splitlines()

            all_projects = []
            for i, url in enumerate(project_links):
                print(f"[{i+1}/{len(project_links)}] {url}")
                project = scrape_project(url)
                if project:
                    all_projects.append(project)
                time.sleep(2)

            ciudades_california = [
                "California", "Los Angeles", "San Francisco", "San Diego",
                "Sacramento", "San Jose", "Oakland", "Santa Monica", "Berkeley",
                "Pasadena", "Long Beach", "Santa Barbara", "Fresno", "Anaheim",
                "Irvine", "Malibu", "Beverly Hills", "Palm Springs", "San Luis Obispo",
                "Redwood City", "Palo Alto", "Glendale", "Burbank", "Riverside", "Ventura",
                "Modesto", "Stockton", "Huntington Beach", "Chula Vista", "Oxnard"
            ]

            def cumple_condiciones(proyecto):
                try:
                    año_valido = int(proyecto["Año"]) >= 2021 if proyecto["Año"] != "No disponible" else False
                except:
                    año_valido = False
                ubicacion_valida = any(ciudad.lower() in proyecto["Ubicación"].lower() for ciudad in ciudades_california)
                return año_valido and ubicacion_valida

            proyectos_filtrados = [p for p in all_projects if cumple_condiciones(p)]

            if proyectos_filtrados:
                file_name = f"proyectos_{categoria}_california_2021.xlsx"
                df = pd.DataFrame(proyectos_filtrados)
                df.to_excel(file_name, index=False)
                print(f"✅ {len(proyectos_filtrados)} proyectos guardados en '{file_name}'")
                customize_excel(file_name)
            else:
                print(f"⚠️ No se encontraron proyectos válidos en {categoria}")

    except Exception as e:
        print(f"Error en el proceso principal: {e}")

if __name__ == "__main__":
    main()